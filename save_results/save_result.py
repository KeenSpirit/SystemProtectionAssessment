"""
Excel workbook generation for protection assessment results.

This module creates formatted Excel workbooks containing fault study
results, protection reach factors, and conductor damage assessments.
Output files are saved to the user's local directory.

Output Sheets:
    - General Information: Study parameters, grid data, settings
    - Summary Results: One row per device across all feeders
    - Detailed Results: One row per terminal across all devices
    - Cond Dmg Results: One row per line section (if selected)

Functions:
    save_dataframe: Main entry point for Excel output generation
    format_grid_data: Format external grid parameters
    format_fl_results: Format fault level results per feeder
    format_study_results: Format one feeder's device summary rows
    format_open_points: Join a feeder's open point names
    format_detailed_results: Format one feeder's terminal-level rows
"""

from pathlib import Path
import os
import re
import time
from typing import Any, Dict, List, Optional

import logging

logger = logging.getLogger(__name__)

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from pf_config import pft
import assets as ast
from fault_study import fault_impedance
from relays.reach_factors import device_reach_factors
from relays.elements import get_prot_elements
from save_results import cond_dmg_results as cd
from importlib import reload

reload(fault_impedance)
reload(cd)

# =============================================================================
# SHEET LAYOUTS
# =============================================================================

# Column order for the flat Summary Results table. The first two columns
# identify the row; columns 3-23 are the per-device metrics that were
# formerly row labels in the transposed per-feeder blocks; the last column
# carries the feeder's open points, repeated on every row of that feeder.
# Label spacing is preserved verbatim from the previous layout so that
# existing downstream lookups keep matching.
#
# The four coordination columns are always present. They are blank for
# any device prot_coordination could not assess (missing fault level or
# pickup data), and blank for every device if the coordination stage did
# not run.
SUMMARY_COLUMNS = [
    'Feeder',
    'Protection device',
    'L-L Voltage (kV)',
    'No. Phases',
    'DS Capacity  (kVA)',
    'Max 3Ph FL',
    'Max 2Ph FL',
    'Max PG FL',
    'Min 3Ph FL',
    'Min 2Ph FL',
    'Min PG FL',
    'Min SN 2P FL',
    'Min SN PG FL',
    'Max DS TR (Site name)',
    'Max TR size (kVA)',
    'TR Max Ph ',
    'TR Max PG',
    'Downstream Devices',
    'Back-up Device',
    'Ph Coord Margin (s)',
    'Ph Coord FL (A)',
    'PG Coord Margin (s)',
    'PG Coord FL (A)',
    'Feeder Open Points',
]

# Column order for the flat Detailed Results table. 'Feeder' and
# 'Primary Protection' identify the row; 'Termination' replaces the
# device-named column of the old per-device blocks, which is what made
# those blocks impossible to stack.
#
# The NPS columns are always present, even though a device with all NPS
# elements out of service contributes no NPS values (see nps_oos). A
# fixed schema means every project in the fleet produces an identically
# shaped sheet, which a varying one would not; devices without NPS get
# blank cells rather than a missing column.
DETAILED_COLUMNS = [
    'Feeder',
    'Primary Protection',
    'Tfmr Size (kVA)',
    'Termination',
    'Construction',
    'Max 3P fault',
    'Max 2P fault',
    'Max PG fault',
    'Min 3P fault',
    'Min 2P fault',
    'Min PG fault',
    'Min SN 2P fault',
    'Min SN PG fault',
    'EF PRI PU',
    'EF BU PU',
    'PH PRI PU',
    'PH BU PU',
    'NPS PRI PU',
    'NPS BU PU',
    'EF PRI RF',
    'EF BU RF',
    'PH PRI RF',
    'PH BU RF',
    'NPS EF PRI RF',
    'NPS EF BU RF',
    'NPS PH PRI RF',
    'NPS PH BU RF',
]

# =============================================================================
# OUTPUT PATH RESOLUTION
# =============================================================================

def _resolve_output_path(app: pft.Application) -> Path:
    """
    Resolve the output directory for the Excel results file.

    Probes the Citrix client path first and falls back to the local
    path if the probe fails. The probe is guarded against ``OSError``
    to handle cases where ``\\\\client\\c$\\`` is partially reachable
    but unstable — this occurs when Citrix client drive mapping is
    disabled by policy, the user's endpoint (e.g. a VM) is not
    configured to expose its local drive, or the Citrix session
    experiences a transient network fault during the probe.

    When the fallback is triggered by an OSError, a warning is
    surfaced via ``logger.warning`` to alert the user that files will
    be written inside the Citrix session and may not persist after
    logoff.

    Args:
        app: PowerFactory application instance, used for warning
            output.

    Returns:
        Path to the output directory.

    Note:
        Attempts paths in order:
            1. //client/c$/LocalData/{username}/ (Citrix)
            2. c:/LocalData/{username}/ (Local)
    """
    user = Path.home().name
    basepath = Path('//client/c$/LocalData') / user

    try:
        basepath_exists = basepath.exists()
        probe_failed = False
    except OSError as err:
        # Citrix client drive mapping unavailable or network flaky.
        # Covers WinError 53, 64, 67, 1231 and similar transient
        # network errors that Path.exists() does not swallow.
        basepath_exists = False
        probe_failed = True
        logger.warning(
            f"Citrix client path '{basepath}' could not be probed "
            f"({err}). Falling back to local path. If running via "
            f"Citrix, output files will be written inside the Citrix "
            f"session and may not persist after logoff — retrieve "
            f"them before ending the session."
        )

    if basepath_exists:
        return basepath

    local_path = Path('c:/LocalData') / user

    # Only warn about a missing local path when the Citrix probe
    # actually succeeded and reported 'does not exist' (i.e. the
    # user is likely on a local install). If the probe raised, the
    # warning above has already been emitted.
    if not probe_failed and not local_path.exists():
        logger.warning(
            f"Local output path '{local_path}' does not exist. "
            f"The results file may fail to save."
        )

    return local_path

# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def save_dataframe(
    app: pft.Application,
    region: str,
    study_selections: List[str],
    external_grid: Dict,
    feeders: List,
    output_dir: Optional[Path] = None,
) -> str:
    """
    Save protection assessment results to an Excel workbook.

    Creates a formatted Excel file containing all study results.
    The file is saved to the user's local directory with a timestamp.

    Args:
        app: PowerFactory application instance.
        region: Network region ('SEQ' or 'Regional Models').
        study_selections: List of selected study types.
        external_grid: Dictionary of grid objects to parameter lists.
        feeders: List of Feeder dataclasses with populated results.

    Output Location:
        Attempts paths in order:
        1. //client/c$/LocalData/{username}/ (Citrix)
        2. c:/LocalData/{username}/ (Local)

    Example:
        >>> save_dataframe(app, 'SEQ', selections, grids, feeders)
        Output file saved to //client/c$/LocalData/user/Fault Study...
    """
    project = app.GetActiveProject()
    derived_proj = project.der_baseproject

    try:
        der_proj_name = derived_proj.GetFullName()
    except AttributeError:
        der_proj_name = project.loc_name

    try:
        project_version = project.der_baseversion
    except AttributeError:
        project_version = 'NA'

    logger.info("Saving Fault Level Study...")

    # Generate filename with timestamp. The project name (not the
    # study case) provides file identity: in batch runs the study
    # case is 'All Active Grids Study Case' for every project, so
    # it distinguishes nothing. The study case is still recorded on
    # the General Information sheet.
    date_string = time.strftime("%Y%m%d-%H%M%S")
    study_case_name = app.GetActiveStudyCase().loc_name
    filename = fix_string(
        f'Fault Study Results {project.loc_name} {date_string}.xlsx'
    )

    # Output path: injected directory (batch) or the legacy
    # Citrix/local per-user probe (interactive).
    if output_dir is not None:
        clientpath = Path(output_dir)
        clientpath.mkdir(parents=True, exist_ok=True)
    else:
        clientpath = _resolve_output_path(app)

    filepath = os.path.join(clientpath, filename)

    # Format data for output
    formatted_grid_data = format_grid_data(external_grid)
    grid_data_df = pd.DataFrame(formatted_grid_data)
    grid_data_df = clean_dataframe(grid_data_df)

    fault_studies_pd = format_fl_results(region, feeders)

    # Regional fault impedance values
    if region == 'SEQ':
        oh_z = '0'
        ug_z = '0'
    else:
        oh_z = '50'
        ug_z = '10'

    variations = app.GetActiveNetworkVariations()

    # Write to Excel
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        workbook = writer.book

        # General Information sheet
        grid_data_df.to_excel(
            writer,
            sheet_name='General Information',
            startrow=26,
            index=False
        )
        worksheet = workbook['General Information']

        _write_general_info(
            worksheet, study_case_name, date_string, der_proj_name,
            project_version, oh_z, ug_z, variations
        )

        # Summary Results sheet: every feeder's devices stacked into one
        # table, written in a single call so the sheet is one contiguous
        # range that post-processing can read as a table.
        # Empty frames must be excluded: an all-object empty frame wins
        # the dtype negotiation in pd.concat and turns every numeric
        # column into object, which pandas 3.x does not warn about.
        summary_frames = [
            key[0] for key in fault_studies_pd.values()
            if not key[0].empty
        ]

        if summary_frames:
            summary_df = pd.concat(summary_frames, ignore_index=True)
        else:
            summary_df = pd.DataFrame(columns=SUMMARY_COLUMNS)

        summary_df = clean_dataframe(summary_df)
        summary_df = ensure_numeric_types(summary_df)
        summary_df.to_excel(
            writer,
            sheet_name='Summary Results',
            startrow=0,
            index=False
        )

        # Detailed Results sheet: every device on every feeder stacked
        # into one table, written in a single call.
        detailed_frames = [
            key[1] for key in fault_studies_pd.values()
            if not key[1].empty
        ]

        if detailed_frames:
            detailed_df = pd.concat(detailed_frames, ignore_index=True)
        else:
            detailed_df = pd.DataFrame(columns=DETAILED_COLUMNS)

        detailed_df = clean_dataframe(detailed_df)
        detailed_df = ensure_numeric_types(detailed_df)
        detailed_df.to_excel(
            writer,
            sheet_name='Detailed Results',
            startrow=0,
            index=False
        )

        # Cond Dmg Results sheet: every line section on every feeder
        # stacked into one table.
        if 'Conductor Damage Assessment' in study_selections:
            cond_frames = [
                frame for frame in (
                    cd.cond_damage_results(fdr) for fdr in feeders
                )
                if not frame.empty
            ]

            if cond_frames:
                cond_damage_df = pd.concat(cond_frames, ignore_index=True)
            else:
                cond_damage_df = pd.DataFrame(
                    columns=cd.COND_DAMAGE_COLUMNS
                )

            cond_damage_df = clean_dataframe(cond_damage_df)
            cond_damage_df = ensure_numeric_types(cond_damage_df)
            cond_damage_df.to_excel(
                writer,
                sheet_name='Cond Dmg Results',
                startrow=0,
                index=False
            )

    # Apply formatting
    wb = load_workbook(filepath)
    ws = wb['General Information']
    adjust_gen_info_col_size(ws)

    ws = wb['Summary Results']
    adjust_summ_col_size(ws)

    ws = wb['Detailed Results']
    adjust_detailed_col_size(ws)

    if 'Cond Dmg Results' in wb.sheetnames:
        ws = wb['Cond Dmg Results']
        adjust_cond_damage_col_width(ws)

    wb.save(filepath)
    logger.info(f"Output file saved to {filepath}")
    return filepath


def _write_general_info(
    worksheet,
    study_case: str,
    date_string: str,
    project_name: str,
    version: str,
    oh_z: str,
    ug_z: str,
    variations: List
) -> None:
    """Write general information content to worksheet."""
    safe_set_cell(worksheet, 'A1', str(study_case))
    safe_set_cell(worksheet, 'A2', "Fault Level Study")
    safe_set_cell(worksheet, 'A4', 'Script Run Date-Time')
    safe_set_cell(worksheet, 'A5', str(date_string))
    safe_set_cell(worksheet, 'A6', 'Base Project:')
    safe_set_cell(worksheet, 'A7', str(project_name))
    safe_set_cell(worksheet, 'A8', 'Used Version:')
    safe_set_cell(worksheet, 'A9', str(version))
    safe_set_cell(worksheet, 'A10', 'Network Variations:')

    if variations:
        var_names = ', '.join([v.loc_name for v in variations])
        safe_set_cell(worksheet, 'A11', var_names)
    else:
        safe_set_cell(worksheet, 'A11', 'None')

    safe_set_cell(worksheet, 'A13', 'Short-circuit calculation method:')
    safe_set_cell(worksheet, 'B13', 'Complete')
    safe_set_cell(worksheet, 'A14', 'Maximum fault c-factor:')
    safe_set_cell(worksheet, 'B14', '1.1')
    safe_set_cell(worksheet, 'A15', 'Minimum fault c-factor:')
    safe_set_cell(worksheet, 'B15', '1.0')
    safe_set_cell(worksheet, 'A17', 'OH Line Minimum Earth Fault Impedance:')
    safe_set_cell(worksheet, 'B17', f'{oh_z} ohms')
    safe_set_cell(worksheet, 'A18', 'UG Cable Minimum Earth Fault Impedance:')
    safe_set_cell(worksheet, 'B18', f'{ug_z} ohms')
    safe_set_cell(worksheet, 'A20', 'Reach Factor Thresholds:')
    safe_set_cell(worksheet, 'A21', 'Primary RF >= 2.0 (SEQ) or 1.7 (Regional)')
    safe_set_cell(worksheet, 'A22', 'Backup RF >= 1.3')
    safe_set_cell(
        worksheet,
        'A23',
        'Tabulated fault clearing time shown is for final trip.'
    )
    safe_set_cell(worksheet, 'A25', 'External Grid Data:')


# =============================================================================
# DATA FORMATTING FUNCTIONS
# =============================================================================

def format_grid_data(ext_grid: Dict) -> Dict:
    """
    Format external grid parameters for DataFrame creation.

    Args:
        ext_grid: Dictionary mapping grid objects to parameter lists.

    Returns:
        Dictionary with 'Parameter' column and columns for each grid's
        Maximum, Minimum, and System Normal Minimum values.
    """
    formatted_grid_data = {}

    for grid, attributes in ext_grid.items():
        formatted_grid_data['Parameter'] = [
            '3-P fault level (A):',
            'R/X:',
            'Z2/Z1:',
            'X0/X1:',
            'R0/X0:'
        ]
        formatted_grid_data[f'{grid.loc_name} Maximum'] = attributes[:5]
        formatted_grid_data[f'{grid.loc_name} Minimum'] = attributes[5:10]
        formatted_grid_data[f'{grid.loc_name} Sys Norm Minimum'] = attributes[-5:]

    return formatted_grid_data


def format_fl_results(region: str, feeders: List) -> Dict:
    """
    Format fault level results for all feeders.

    Args:
        region: Network region string.
        feeders: List of Feeder dataclasses.

    Returns:
        Dictionary mapping feeder names to lists containing:
        [summary_df, detailed_df]

        The per-feeder frames share ``SUMMARY_COLUMNS`` and
        ``DETAILED_COLUMNS`` respectively, and are concatenated by
        ``save_dataframe`` into one Summary Results table and one
        Detailed Results table. Open points are carried inside the
        summary frame, so the separate open-points frame no longer
        exists.
    """
    fault_studies_pd = {}

    for feeder in feeders:
        try:
            summary_df = format_study_results(feeder)
            detailed_df = format_detailed_results(region, feeder)
        except Exception:
            logger.exception("Fault level formatting failed for feeder %s; skipping", feeder.loc_name)
            continue
        fault_studies_pd[feeder.obj.loc_name] = [
            summary_df,
            detailed_df
        ]

    return fault_studies_pd


def format_study_results(feeder) -> pd.DataFrame:
    """
    Format one feeder's device summary as rows of the flat table.

    Produces one row per protection device, with the feeder name and
    the feeder's open-point list repeated on every row so the table can
    be looked up or filtered without reference to any other sheet.

    Args:
        feeder: Feeder dataclass with ``obj``, ``devices`` and
            ``open_points`` attributes.

    Returns:
        DataFrame with exactly the ``SUMMARY_COLUMNS`` columns. Numeric
        cells that have no value are left as NaN (not ""), so that the
        concatenated frame keeps numeric dtypes and Excel receives
        genuinely empty cells rather than text.
    """
    feeder_name = str(feeder.obj.loc_name)
    open_points = format_open_points(feeder)

    rows = []

    for device in feeder.devices:
        ds_names = [str(d.obj.loc_name) for d in device.ds_devices]
        us_names = [str(d.obj.loc_name) for d in device.us_devices]

        # max_ds_tr may be absent, or present with no terminal resolved.
        # Both are reported as blanks rather than crashing the run.
        max_ds_tr = getattr(device, 'max_ds_tr', None)
        tr_site = ''
        tr_kva = None
        tr_max_ph = None
        tr_max_pg = None

        if max_ds_tr is not None:
            if getattr(max_ds_tr, 'term', None) is not None:
                tr_site = str(max_ds_tr.term.cpSubstat.loc_name)
            tr_kva = safe_numeric(getattr(max_ds_tr, 'load_kva', None))
            tr_max_ph = safe_numeric(getattr(max_ds_tr, 'max_ph', None))
            tr_max_pg = safe_numeric(getattr(max_ds_tr, 'max_pg', None))

        rows.append({
            'Feeder': feeder_name,
            'Protection device': str(device.obj.loc_name),
            'L-L Voltage (kV)': safe_numeric(device.l_l_volts),
            'No. Phases': safe_numeric(device.phases),
            'DS Capacity  (kVA)': safe_numeric(device.ds_capacity),
            'Max 3Ph FL': safe_numeric(device.max_fl_3ph),
            'Max 2Ph FL': safe_numeric(device.max_fl_2ph),
            'Max PG FL': safe_numeric(device.max_fl_pg),
            'Min 3Ph FL': safe_numeric(device.min_fl_3ph),
            'Min 2Ph FL': safe_numeric(device.min_fl_2ph),
            'Min PG FL': safe_numeric(device.min_fl_pg),
            'Min SN 2P FL': safe_numeric(device.min_sn_fl_2ph),
            'Min SN PG FL': safe_numeric(device.min_sn_fl_pg),
            'Max DS TR (Site name)': tr_site,
            'Max TR size (kVA)': tr_kva,
            'TR Max Ph ': tr_max_ph,
            'TR Max PG': tr_max_pg,
            'Downstream Devices': ', '.join(ds_names),
            'Back-up Device': ', '.join(us_names),
            'Ph Coord Margin (s)': safe_round(device.ph_coord_margin),
            'Ph Coord FL (A)': safe_numeric(device.ph_coord_fl),
            'PG Coord Margin (s)': safe_round(device.pg_coord_margin),
            'PG Coord FL (A)': safe_numeric(device.pg_coord_fl),
            'Feeder Open Points': open_points,
        })

    if not rows:
        # A feeder with no protection devices contributes no device rows.
        # Emit a single identifying row so the feeder — and its open
        # points — are not silently dropped from the table.
        rows.append({
            'Feeder': feeder_name,
            'Protection device': '',
            'Feeder Open Points': open_points,
        })

    return pd.DataFrame(rows, columns=SUMMARY_COLUMNS)


def format_open_points(feeder) -> str:
    """
    Format a feeder's open points as a single comma-separated string.

    The flat Summary Results table has one row per device, so the open
    points can no longer occupy their own column of names; they are
    collapsed into one cell and repeated on each of the feeder's rows.

    Args:
        feeder: Feeder dataclass with an ``open_points`` attribute.

    Returns:
        Comma-separated open point names, or '' if there are none.
    """
    open_points = getattr(feeder, 'open_points', None) or []
    return ', '.join(str(op.loc_name) for op in open_points)


def format_detailed_results(region: str, feeder) -> pd.DataFrame:
    """
    Format one feeder's terminal-level results as rows of the flat table.

    Each of the feeder's devices contributes one row per terminal in its
    protection section, tagged with the feeder name and the device name,
    so that all devices on all feeders stack into a single table.

    Rows remain sorted by 'Max PG fault' descending within each device,
    and devices appear in ``feeder.devices`` order.

    Args:
        region: Network region string ('SEQ' or 'Regional Models').
        feeder: Feeder dataclass with ``obj`` and ``devices``.

    Returns:
        DataFrame with exactly the ``DETAILED_COLUMNS`` columns. Devices
        with no in-service NPS elements leave the NPS columns blank.
    """
    feeder_name = str(feeder.obj.loc_name)
    frames = []

    for device in feeder.devices:
        device_name = str(device.obj.loc_name)
        elements = device.sect_terms
        count = len(elements)

        # Reach factors are calculated in the pipeline (see
        # populate_reach_factors); this layer renders them. The stored
        # lists are positionally aligned with sect_terms as it stood at
        # calculation time, so verify that order still holds before
        # using them - a silent misalignment would attach every reach
        # factor to the wrong terminal.
        stored_terms = device.reach_factor_terms
        aligned = (
            device.reach_factors is not None
            and len(stored_terms) == count
            and all(a is b for a, b in zip(stored_terms, elements))
        )

        if aligned:
            dev_reach_factors = device.reach_factors
        elif device.reach_factors is None:
            logger.warning(
                "No stored reach factors for %s/%s; recalculating. The "
                "reach factor stage did not run for this feeder.",
                feeder_name, device_name
            )
            dev_reach_factors = device_reach_factors(region, device, elements)
        else:
            logger.warning(
                "Stored reach factors for %s/%s do not match the current "
                "sect_terms order; recalculating. Something re-ordered "
                "sect_terms after the reach factor stage.",
                feeder_name, device_name
            )
            dev_reach_factors = device_reach_factors(region, device, elements)

        fault_levels = {
            'Tfmr Size (kVA)': [None] * count,
            'Termination': [str(e.obj.loc_name) for e in elements],
            'Construction': [e.constr for e in elements],
            'Max 3P fault': [safe_numeric(e.max_fl_3ph) for e in elements],
            'Max 2P fault': [safe_numeric(e.max_fl_2ph) for e in elements],
            'Max PG fault': [safe_numeric(e.max_fl_pg) for e in elements],
            'Min 3P fault': [safe_numeric(e.min_fl_3ph) for e in elements],
            'Min 2P fault': [safe_numeric(e.min_fl_2ph) for e in elements],
            'Min PG fault': [safe_numeric(e.min_fl_pg) for e in elements],
            'Min SN 2P fault': [safe_numeric(e.min_sn_fl_2ph) for e in elements],
            'Min SN PG fault': [safe_numeric(e.min_sn_fl_pg) for e in elements]
        }

        pick_ups = {
            'EF PRI PU': _padded(dev_reach_factors.get('ef_pickup'), count),
            'EF BU PU': _padded(dev_reach_factors.get('bu_ef_pickup'), count),
            'PH PRI PU': _padded(dev_reach_factors.get('ph_pickup'), count),
            'PH BU PU': _padded(dev_reach_factors.get('bu_ph_pickup'), count)
        }

        reach_factors = {
            'EF PRI RF': _padded(dev_reach_factors.get('ef_rf'), count),
            'EF BU RF': _padded(dev_reach_factors.get('bu_ef_rf'), count),
            'PH PRI RF': _padded(dev_reach_factors.get('ph_rf'), count),
            'PH BU RF': _padded(dev_reach_factors.get('bu_ph_rf'), count),
        }

        # Nps results are only populated if there are nps elements in
        # service. The NPS columns exist either way; a device with none
        # in service simply leaves them blank.
        if not nps_oos(device):
            pick_ups.update({
                'NPS PRI PU': _padded(
                    dev_reach_factors.get('nps_pickup'), count
                ),
                'NPS BU PU': _padded(
                    dev_reach_factors.get('bu_nps_pickup'), count
                ),
            })
            reach_factors.update({
                'NPS EF PRI RF': _padded(
                    dev_reach_factors.get('nps_ef_rf'), count
                ),
                'NPS EF BU RF': _padded(
                    dev_reach_factors.get('bu_nps_ef_rf'), count
                ),
                'NPS PH PRI RF': _padded(
                    dev_reach_factors.get('nps_ph_rf'), count
                ),
                'NPS PH BU RF': _padded(
                    dev_reach_factors.get('bu_nps_ph_rf'), count
                ),
            })

        df = pd.DataFrame(fault_levels | pick_ups | reach_factors)

        # Sort by Max PG fault descending
        if 'Max PG fault' in df.columns and not df.empty:
            try:
                df = df.sort_values(by='Max PG fault', ascending=False)
            except (AttributeError, KeyError):
                pass

        # Map transformer sizes. Keyed off 'Termination', which now
        # carries the terminal names formerly held in the device-named
        # column.
        try:
            tr_dict = {
                str(load.term.loc_name): safe_numeric(load.load_kva)
                for load in device.sect_loads
                if hasattr(load, 'term') and hasattr(load, 'load_kva')
            }
            df['Tfmr Size (kVA)'] = df['Termination'].map(tr_dict)
        except (AttributeError, KeyError):
            pass

        df.insert(0, 'Primary Protection', device_name)
        df.insert(0, 'Feeder', feeder_name)

        # Surface anything device_reach_factors starts returning that the
        # schema does not know about, rather than dropping it silently.
        unexpected = [c for c in df.columns if c not in DETAILED_COLUMNS]
        if unexpected:
            logger.warning(
                "Detailed results for %s/%s produced unrecognised "
                "columns which will not be written: %s",
                feeder_name, device_name, unexpected
            )

        frames.append(df.reindex(columns=DETAILED_COLUMNS))

    frames = [f for f in frames if f is not None and not f.empty]
    if not frames:
        return pd.DataFrame(columns=DETAILED_COLUMNS)

    return pd.concat(frames, ignore_index=True)


# =============================================================================
# DATA CLEANING FUNCTIONS
# =============================================================================

def safe_numeric(value: Any) -> Any:
    """
    Safely convert value to numeric, preserving None/NaN.

    Args:
        value: Value to convert.

    Returns:
        Numeric value or None if conversion fails.
    """
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    try:
        return float(value) if value else None
    except (ValueError, TypeError):
        return None


def safe_round(value: Any, digits: int = 3) -> Any:
    """
    Coerce a value to a number and round it for display.

    Rounding happens here rather than in the assessment modules so the
    stored dataclass values keep full precision and only the Excel
    presentation is truncated.

    Args:
        value: Raw value from a dataclass attribute.
        digits: Decimal places to round to.

    Returns:
        The rounded float, or whatever ``safe_numeric`` produced for a
        missing value (left as-is so the column keeps a numeric dtype
        and Excel receives an empty cell).
    """
    numeric = safe_numeric(value)
    if numeric is None or pd.isna(numeric):
        return numeric
    return round(float(numeric), digits)


def _padded(values: Optional[List], length: int) -> List:
    """
    Return ``values`` resized to ``length``, padding with None.

    ``device_reach_factors`` returns a list per terminal, but a missing
    or short key would previously make ``pd.DataFrame`` raise "All
    arrays must be of the same length" and take the whole project's
    output with it. Padding with None keeps the row count honest: a
    blank cell means no value was produced, which is what happened.

    Args:
        values: List from device_reach_factors, or None if absent.
        length: Number of terminals in the device's section.

    Returns:
        List of exactly ``length`` items.
    """
    if not values:
        return [None] * length

    if len(values) < length:
        return list(values) + [None] * (length - len(values))

    if len(values) > length:
        logger.warning(
            "Reach factor list of %s values truncated to %s terminals; "
            "device_reach_factors returned more values than elements.",
            len(values), length
        )

    return list(values[:length])


def clean_string_value(value: Any) -> str:
    """
    Clean string values to remove problematic characters.

    Removes control characters and limits length for Excel compatibility.

    Args:
        value: Value to clean.

    Returns:
        Cleaned string suitable for Excel cells.
    """
    if pd.isna(value) or value is None:
        return ''

    value_str = str(value)

    # Remove control characters (ASCII 0-31 except tab, newline, CR)
    value_str = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', value_str)

    # Replace specific problematic characters
    replacements = {
        '\x00': '', '\x01': '', '\x02': '', '\x03': '',
        '\x04': '', '\x05': '', '\x06': '', '\x07': '',
        '\x08': '', '\x0B': ' ', '\x0C': ' ', '\x0E': '',
        '\x0F': ''
    }

    for old, new in replacements.items():
        value_str = value_str.replace(old, new)

    # Limit to Excel cell character limit
    if len(value_str) > 32767:
        value_str = value_str[:32767]

    return value_str


def safe_set_cell(worksheet, cell_reference: str, value: Any) -> None:
    """
    Safely set cell value with proper cleaning.

    Args:
        worksheet: openpyxl worksheet object.
        cell_reference: Cell reference string (e.g., 'A1').
        value: Value to set.
    """
    try:
        cleaned_value = clean_string_value(value)
        worksheet[cell_reference] = cleaned_value
    except Exception:
        worksheet[cell_reference] = str(value)[:32767] if value else ''


def create_safe_sheet_name(name: str) -> str:
    """
    Create a valid Excel sheet name.

    Excel sheet names have restrictions:
    - Max 31 characters
    - Cannot contain: \\ / ? * [ ] :
    - Cannot be empty
    - Cannot start/end with single quote

    Args:
        name: Proposed sheet name.

    Returns:
        Valid Excel sheet name.
    """
    if not name:
        name = "Sheet"

    name_str = clean_string_value(str(name))

    forbidden_chars = ['\\', '/', '?', '*', '[', ']', ':']
    for char in forbidden_chars:
        name_str = name_str.replace(char, '_')

    name_str = name_str.strip("'")

    if not name_str or name_str.isspace():
        name_str = "Sheet"

    if len(name_str) > 31:
        name_str = name_str[:31]

    name_str = name_str.rstrip()

    return name_str


def fix_string(file_name: str) -> str:
    """
    Remove invalid characters from filename.

    Args:
        file_name: Proposed filename.

    Returns:
        Valid filename with forbidden characters replaced.
    """
    if not file_name:
        return "default_filename"

    file_name_str = str(file_name)
    forbidden_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']

    for char in forbidden_chars:
        file_name_str = file_name_str.replace(char, '_')

    file_name_str = re.sub(r'[\x00-\x1F\x7F]', '_', file_name_str)

    return file_name_str


def ensure_numeric_types(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure numeric values are stored as proper numeric types.

    Converts string representations of numbers to actual numeric types
    for proper Excel formatting.

    Args:
        df: DataFrame to process.

    Returns:
        DataFrame with proper numeric types.
    """
    if df is None or df.empty:
        return df

    df_numeric = df.copy()

    for col in df_numeric.columns:
        col_str = str(col).lower()

        # Skip string columns
        if any(
                kw in col_str
                for kw in [
                    'name', 'construction', 'site', 'device', 'feeder',
                    'point', 'termination', 'protection'
                ]
        ):
            continue

        # Attempt numeric conversion; leave the column untouched if it
        # cannot be converted.
        try:
            df_numeric[col] = pd.to_numeric(df_numeric[col])
        except (ValueError, TypeError):
            continue

        # Convert whole numbers to integers
        if df_numeric[col].dtype in ['float64', 'float32']:
            non_null = df_numeric[col].dropna()
            if len(non_null) > 0:
                all_whole = all(
                    pd.isna(val) or (isinstance(val, (int, float)) and val == int(val))
                    for val in non_null
                )
                if all_whole:
                    df_numeric[col] = df_numeric[col].astype('Int64')

    return df_numeric


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean DataFrame for Excel compatibility.

    Replaces inf values and cleans string columns.

    Args:
        df: DataFrame to clean.

    Returns:
        Cleaned DataFrame.
    """
    if df is None or df.empty:
        return df

    df_clean = df.copy()
    df_clean = df_clean.replace([float('inf'), float('-inf')], None)

    for col in df_clean.columns:
        if df_clean[col].dtype == 'object':
            df_clean[col] = df_clean[col].apply(
                lambda x: clean_string_value(x) if isinstance(x, str) else x
            )

    df_clean.columns = [clean_string_value(str(col)) for col in df_clean.columns]

    return df_clean


# =============================================================================
# COLUMN WIDTH ADJUSTMENT FUNCTIONS
# =============================================================================

def adjust_gen_info_col_size(ws) -> None:
    """Adjust column widths for General Information sheet."""
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"].font = Font(bold=True, size=12)

    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 18.0

    for cell in ws[27]:
        cell.alignment = Alignment(wrap_text=True)


def adjust_summ_col_size(ws) -> None:
    """
    Format the flat Summary Results table.

    Bolds and wraps the single header row, freezes the header plus the
    two identifying columns, applies an autofilter over the used range,
    and sizes each column to its content within sensible bounds.
    """
    if ws.max_row < 1 or ws.max_column < 1:
        return

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='bottom')

    ws.row_dimensions[1].height = 30.0

    # Freeze the header row and the Feeder / Protection device columns.
    ws.freeze_panes = 'C2'

    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f'A1:{last_col}{ws.max_row}'

    for col in ws.columns:
        column = col[0].column_letter
        max_length = 0

        for cell in col:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (AttributeError, TypeError):
                pass

        # Clamp so the open-points and downstream-device columns, which
        # can hold very long joined lists, do not push the table off
        # screen.
        ws.column_dimensions[column].width = min(max(max_length + 2, 9), 70)


def adjust_detailed_col_size(ws) -> None:
    """
    Format the flat Detailed Results table.

    Bolds and wraps the single header row, freezes the header plus the
    two identifying columns, applies an autofilter over the used range,
    and sizes each column to its content within sensible bounds.
    """
    if ws.max_row < 1 or ws.max_column < 1:
        return

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='bottom')

    ws.row_dimensions[1].height = 30.0

    # Freeze the header row and the Feeder / Primary Protection columns.
    ws.freeze_panes = 'C2'

    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f'A1:{last_col}{ws.max_row}'

    for col in ws.columns:
        column = col[0].column_letter
        max_length = 0

        for cell in col:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (AttributeError, TypeError):
                pass

        ws.column_dimensions[column].width = min(max(max_length + 2, 9), 45)


def adjust_cond_damage_col_width(ws) -> None:
    """
    Format the flat Cond Dmg Results table.

    Bolds and wraps the single header row, freezes the header plus the
    two identifying columns, applies an autofilter over the used range,
    and sizes each column to its content within sensible bounds.
    """
    if ws.max_row < 1 or ws.max_column < 1:
        return

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='bottom')

    ws.row_dimensions[1].height = 30.0

    # Freeze the header row and the Feeder / Device columns.
    ws.freeze_panes = 'C2'

    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f'A1:{last_col}{ws.max_row}'

    for col in ws.columns:
        column = col[0].column_letter
        max_length = 0

        for cell in col:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (AttributeError, TypeError):
                pass

        ws.column_dimensions[column].width = min(max(max_length + 2, 9), 45)


def nps_oos(device: ast.Device) -> bool:
    """
    Determine whether any nps elements related to the device
    (including upstream devices) are in service.
    Returns True if all nps elements are out of service.
    Returns False if at least one nps element is in service.
    """
    section_devices = [device.obj]
    section_devices.extend([bu_device.obj for bu_device in device.us_devices])
    all_elements = [get_prot_elements(device_pf) for device_pf in section_devices
                    if device_pf.GetClassName() == 'ElmRelay']
    if not all_elements:
        return True
    all_elements = all_elements[0]
    nps_elements = (
            all_elements['nps_idmt_elements'] + all_elements['nps_inst_elements']
    )
    nps_disabled = all([element.IsOutOfService() for element in nps_elements])
    return nps_disabled