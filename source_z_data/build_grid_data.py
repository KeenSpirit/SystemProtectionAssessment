#!/usr/bin/env python3
"""
build_grid_results.py
=====================

Populate "grid_results_all.xlsx" for the Ergon external grids.

For every unique PowerFactory external grid in `pf_external_grids.xlsx`
(column B), this script finds the matching reporting bus(es) in the
Ergon fault level report and writes three rows:

    Bound = Max     / Scenario = Maximum        <- 'Max-Max Fault Level Report' tab
    Bound = Min     / Scenario = System Normal  <- 'Min Fault Level Report'     tab
    Bound = Min_SN  / Scenario = System Normal  <- duplicate of the Min row

Columns F..J (3P fault, R/X, Z2/Z1, X0/X1, R0/X1) are copied verbatim from
report columns V..Z.

MATCHING RULES (in order of precedence)
---------------------------------------
1. Sub-code.   The first four characters of the grid name are matched against
               the report's 'Sub Code' (column A).  No fuzzy matching is used:
               a code that is absent from the report is reported as unmatched.
2. Voltage.    If the grid name carries a "<n>kV" token it must equal the bus
               'Voltage (kV)' (column E).
3. Identifier. Numbers in the tail of the grid name (T1, TR2, BUS1_BUS2, ...)
               are matched against numbers in the tail of the bus name
               (TEF T1, BUS 2, ...).  Applied only when it yields a non-empty
               candidate set.
4. Fault level.Remaining candidates are grouped by identical Ik"max in the
               'Max Fault Level Report' tab.  If more than one group survives,
               the group is chosen that minimises

                   |V_bus - V_grid| / max(...)  +  0.5 * |W_bus - RX_grid| / max(...)

               where V_grid is the grid's Bound=Max 3P fault (A, converted to
               kA) and RX_grid its R/X ratio.  R/X carries half weight.

Grids whose source data is the PowerFactory default "infinite" external grid
(3P fault = 524864 A, R/X = 0.1) carry no usable fault-level information, so
rule 4 cannot be applied to them.  Where such a grid still has more than one
candidate group, no bus is assigned and the grid is listed in 'Mapping Notes'
with its full candidate list for manual selection.

UNMATCHED GRIDS
---------------
Where no reporting bus is assigned (no sub-code match, no voltage match, or an
unresolved ambiguity), columns C..H of `pf_external_grids.xlsx` are carried
across verbatim instead:

    pf column C (Bound)   -> output column D
    pf column D (3P fault)-> output column F
    pf column E (R/X)     -> output column G
    pf column F (Z2/Z1)   -> output column H
    pf column G (X0/X1)   -> output column I
    pf column H (R0/X1)   -> output column J

The grid's Bound=Max row supplies output row 1 and its Bound=Min row supplies
row 2; row 3 repeats the Min row with the Bound relabelled "Min_SN".  Columns C
(Buses) and E (Scenario) stay empty for these grids.

NOTE ON UNITS: `pf_external_grids.xlsx` reports 3P fault in amps, whereas the
fault level report - and therefore every matched row - is in kA.  The carried
over values are copied verbatim, so column F mixes the two units.  Set
CONVERT_CARRIED_3P_TO_KA = True below to divide the carried values by 1000.

Usage
-----
    python build_grid_results.py \
        --pf        pf_external_grids.xlsx \
        --report    "2026 Fault Level Report Ergon  Internal_V1_1.xlsx" \
        --template  grid_results_all.xlsx \
        --out       grid_results_all_completed.xlsx
"""

from __future__ import annotations

import argparse
import re
import shutil
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font

# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #

MAX_MATCH_TAB = "Max Fault Level Report"       # used ONLY to choose the bus
MAX_VALUE_TAB = "Max-Max Fault Level Report"   # source of the Bound=Max values
MIN_VALUE_TAB = "Min Fault Level Report"       # source of the Bound=Min values

# Report column letters copied into output columns F..J
VALUE_COLS = ["V", "W", "X", "Y", "Z"]

# PowerFactory default ("infinite") external grid signature
PLACEHOLDER_3P = 524864.0
PLACEHOLDER_TOL = 1.0

RX_WEIGHT = 0.5          # R/X carries half the weight of the 3P fault term
HIGH_CONF_GAP = 0.05     # score gap to runner-up needed for "High" confidence

# pf_external_grids reports 3P fault in amps; the report is in kA.  Set True to
# divide carried-over values by 1000 so column F is kA throughout.
CONVERT_CARRIED_3P_TO_KA = False

# pf columns carried across for grids with no matching reporting bus
PF_CARRY_COLS = ["F3P", "RX", "Z2Z1", "X0X1", "R0X1"]

BOUND_ROWS = [
    (" Max", "Maximum", "max"),
    (" Min", "System Normal", "min"),
    (" Min_SN", "System Normal", "min"),
]

REPORT_COLS = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ") + ["AA", "AB"]
VOLT_RE = re.compile(r"(\d+(?:\.\d+)?)\s*kV", re.I)
ID_RE = re.compile(r"(\d+)")


# --------------------------------------------------------------------------- #
# Loading
# --------------------------------------------------------------------------- #

def load_report(path: str | Path, tab: str) -> pd.DataFrame:
    """Read one fault level report tab, keyed on the PowerFactory bus name."""
    df = pd.read_excel(path, sheet_name=tab, header=0).iloc[:, : len(REPORT_COLS)]
    df.columns = REPORT_COLS
    df = df[df["D"].notna()].copy()
    df["D"] = df["D"].astype(str).str.strip()
    # The Max-Max tab repeats a handful of buses verbatim; keep the first.
    df = df.drop_duplicates(subset="D").reset_index(drop=True)
    df["A"] = df["A"].astype(str).str.strip()
    df["code"] = df["A"].str.upper().str[:4]
    df["kv"] = pd.to_numeric(df["E"], errors="coerce")
    for c in VALUE_COLS:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def load_pf(path: str | Path, sheet: str = "Sheet3") -> pd.DataFrame:
    """Read the PowerFactory external grid export."""
    pf = pd.read_excel(path, sheet_name=sheet, header=0).iloc[:, :8]
    pf.columns = ["BSP", "Grid", "Bound", "F3P", "RX", "Z2Z1", "X0X1", "R0X1"]
    pf["BSP_raw"] = pf["BSP"].astype(str)
    pf["Grid_raw"] = pf["Grid"].astype(str)
    pf["Bound_raw"] = pf["Bound"].astype(str)
    for c in ("BSP", "Grid", "Bound"):
        pf[c] = pf[c].astype(str).str.strip()
    pf["F3P"] = pd.to_numeric(pf["F3P"], errors="coerce")
    pf["RX"] = pd.to_numeric(pf["RX"], errors="coerce")
    return pf


# --------------------------------------------------------------------------- #
# Name parsing
# --------------------------------------------------------------------------- #

def parse_grid(name: str) -> tuple[str | None, float | None]:
    """Return (four-character sub-code, voltage in kV) parsed from a grid name."""
    upper = name.strip().upper()
    token = re.split(r"[_\s]", upper)[0]
    code = token[:4] if len(token) >= 4 and token[:4].isalnum() else None
    match = VOLT_RE.search(upper)
    return code, (float(match.group(1)) if match else None)


def _tail(name: str, code: str | None) -> str:
    """Everything after the sub-code and the voltage token."""
    upper = name.strip().upper()
    if code and upper.startswith(code):
        upper = upper[len(code):]
    upper = VOLT_RE.sub(" ", upper)
    upper = re.sub(r"\bEXTERNAL GRID\b.*$", " ", upper)
    return upper.strip(" _-")


def ident_set(name: str, code: str | None) -> set[int]:
    """Small integers (transformer / bus numbers) in the tail of a name."""
    return {int(n) for n in ID_RE.findall(_tail(name, code)) if len(n) <= 2}


def is_placeholder(f3p, rx) -> bool:
    """True when the grid still holds PowerFactory's default infinite source."""
    if f3p is None or pd.isna(f3p):
        return True
    f3p = float(f3p)
    return (
        abs(f3p - PLACEHOLDER_3P) < PLACEHOLDER_TOL
        or f3p <= 0.0
        or f3p > 1.0e6
    )


def rel_diff(a: float, b: float) -> float:
    """Symmetric relative difference, 0 when both are zero."""
    a, b = float(a), float(b)
    scale = max(abs(a), abs(b))
    return abs(a - b) / scale if scale > 0 else 0.0


# --------------------------------------------------------------------------- #
# Matching
# --------------------------------------------------------------------------- #

def match_grid(grid: str, mx: pd.DataFrame, f3p, rx) -> dict:
    """Resolve one external grid to a set of reporting buses."""
    out = {
        "buses": [],
        "status": "",
        "confidence": "",
        "score": None,
        "gap": None,
        "candidates": [],
        "note": "",
    }
    code, kv = parse_grid(grid)

    if code is None:
        out["status"] = "unmatched"
        out["note"] = "No four-character sub-code in the grid name."
        return out

    cand = mx[mx["code"] == code]
    if cand.empty:
        out["status"] = "unmatched"
        out["note"] = f"Sub-code {code} does not appear in the fault level report."
        return out

    voltage_confirmed = False
    if kv is not None:
        by_kv = cand[np.isclose(cand["kv"], kv, atol=1e-6)]
        if by_kv.empty:
            available = sorted({v for v in cand["kv"].dropna()})
            out["status"] = "unmatched"
            out["note"] = (
                f"Sub-code {code} found but no {kv:g} kV bus "
                f"(report has {', '.join(f'{v:g}' for v in available)} kV)."
            )
            return out
        cand = by_kv
        voltage_confirmed = True

    out["candidates"] = list(cand["D"])

    notes = []
    ids = ident_set(grid, code)
    if ids:
        by_id = cand[[bool(ids & ident_set(d, code)) for d in cand["D"]]]
        if not by_id.empty and len(by_id) < len(cand):
            cand = by_id
            notes.append("bus selected on matching T/BUS number")

    groups = [(v, sub) for v, sub in cand.groupby("V", sort=True)]

    if len(groups) == 1:
        winner = groups[0][1]
        out["status"] = "matched"
        out["confidence"] = "High" if voltage_confirmed else "Medium"
        if not voltage_confirmed:
            notes.append("no kV token in grid name; all buses at this substation share one fault level")
    elif is_placeholder(f3p, rx):
        out["status"] = "ambiguous"
        out["confidence"] = "Low"
        out["note"] = (
            f"{len(groups)} candidate fault levels at {code} and the grid still holds "
            "the PowerFactory default source (524864 A, R/X 0.1), so the fault level "
            "cannot discriminate. Manual selection required."
        )
        return out
    else:
        scored = []
        for v, sub in groups:
            score = rel_diff(v, float(f3p) / 1000.0) + RX_WEIGHT * rel_diff(
                float(sub["W"].iloc[0]), float(rx)
            )
            scored.append((round(score, 9), float(v), sub))
        scored.sort(key=lambda t: (t[0], t[1]))          # stable, deterministic
        winner = scored[0][2]
        out["score"] = scored[0][0]
        out["gap"] = round(scored[1][0] - scored[0][0], 6)
        out["status"] = "matched"
        out["confidence"] = "High" if out["gap"] > HIGH_CONF_GAP else "Medium"
        notes.append(
            f"chosen from {len(groups)} candidate fault levels on closest "
            f"3P fault + R/X (score {out['score']:.4f}, next {scored[1][0]:.4f})"
        )

    out["buses"] = sorted(winner["D"])
    out["note"] = "; ".join(notes)
    return out


# --------------------------------------------------------------------------- #
# Row construction
# --------------------------------------------------------------------------- #

def value_row(tab: pd.DataFrame, buses: list[str], pick: str) -> tuple[list, str | None]:
    """Return columns V..Z for the chosen bus, plus a warning if buses disagree."""
    sub = tab[tab["D"].isin(buses)]
    if sub.empty:
        return [None] * len(VALUE_COLS), "selected buses not present in this tab"
    warn = None
    if sub["V"].nunique() > 1:
        warn = (
            f"selected buses differ in this tab "
            f"({', '.join(f'{v:.5f}' for v in sorted(sub['V'].unique()))} kA); "
            f"the {'highest' if pick == 'max' else 'lowest'} value was used"
        )
    idx = sub["V"].idxmax() if pick == "max" else sub["V"].idxmin()
    return [sub.loc[idx, c] for c in VALUE_COLS], warn


def flag_shared_buses(records: list[dict]) -> None:
    """Flag substations where several grids landed on the same bus set while
    other candidate buses at that substation and voltage went unclaimed."""
    by_key: dict[tuple, list[dict]] = {}
    for r in records:
        if r["assigned"]:
            by_key.setdefault(r["key"], []).append(r)

    for group in by_key.values():
        if len(group) < 2:
            continue
        candidates = set()
        for r in group:
            candidates |= set(r["candidates"])
        claimed = set()
        for r in group:
            claimed |= set(r["assigned"])
        unclaimed = candidates - claimed
        if not unclaimed:
            continue
        seen: dict[tuple, list[dict]] = {}
        for r in group:
            seen.setdefault(r["assigned"], []).append(r)
        for assigned, shared in seen.items():
            if len(shared) < 2:
                continue
            names = ", ".join(sorted(r["Grid"].strip() for r in shared))
            msg = (
                f"{len(shared)} grids ({names}) resolved to the same bus set while "
                f"{', '.join(sorted(unclaimed))} remained unassigned - worth a manual check"
            )
            for r in shared:
                r["Note"] = "; ".join([n for n in (r["Note"], msg) if n])
                if r["Confidence"] == "High":
                    r["Confidence"] = "Medium"
                r["flag"] = True


def carried_values(src) -> list:
    """Columns D..H of a pf_external_grids row, for output columns F..J."""
    if src is None:
        return [None] * len(PF_CARRY_COLS)
    vals = [None if pd.isna(src[c]) else src[c] for c in PF_CARRY_COLS]
    if CONVERT_CARRIED_3P_TO_KA and vals[0] is not None:
        vals[0] = float(vals[0]) / 1000.0
    return vals


def carried_bound(src, suffix: str = "") -> str | None:
    """Column C of a pf_external_grids row, optionally relabelled."""
    if src is None:
        return None
    return f"{str(src['Bound_raw']).rstrip()}{suffix}"


def build(pf_path, report_path, template_path, out_path) -> dict:
    pf = load_pf(pf_path)
    mx = load_report(report_path, MAX_MATCH_TAB)
    mm = load_report(report_path, MAX_VALUE_TAB)
    mn = load_report(report_path, MIN_VALUE_TAB)

    # One entry per unique grid, in order of first appearance, with the
    # Bulk Supply Point of that first appearance.
    first = pf.drop_duplicates(subset="Grid", keep="first")
    max_rows = pf[pf["Bound"].str.upper() == "MAX"].drop_duplicates("Grid").set_index("Grid")
    min_rows = pf[pf["Bound"].str.upper() == "MIN"].drop_duplicates("Grid").set_index("Grid")

    results, notes, records = [], [], []
    counts = {"matched": 0, "ambiguous": 0, "unmatched": 0}

    for _, g in first.iterrows():
        grid = g["Grid"]
        src = max_rows.loc[grid] if grid in max_rows.index else None
        f3p = None if src is None else src["F3P"]
        rx = None if src is None else src["RX"]

        m = match_grid(grid, mx, f3p, rx)
        counts[m["status"]] += 1

        buses = m["buses"]
        bus_text = ", ".join(buses) if buses else None
        warnings = []

        if buses:
            vals_max, w1 = value_row(mm, buses, "max")
            vals_min, w2 = value_row(mn, buses, "min")
            warnings = [w for w in (w1, w2) if w]
            for bound, scenario, which in BOUND_ROWS:
                vals = vals_max if which == "max" else vals_min
                results.append(
                    [g["BSP_raw"], g["Grid_raw"], bus_text, bound, scenario, *vals]
                )
        else:
            # No reporting bus: carry pf_external_grids columns C..H across.
            snk = min_rows.loc[grid] if grid in min_rows.index else None
            vals_max, vals_min = carried_values(src), carried_values(snk)
            carried = [
                (carried_bound(src) or BOUND_ROWS[0][0], vals_max),
                (carried_bound(snk) or BOUND_ROWS[1][0], vals_min),
                (carried_bound(snk, "_SN") or BOUND_ROWS[2][0], vals_min),
            ]
            for bound, vals in carried:
                results.append([g["BSP_raw"], g["Grid_raw"], None, bound, None, *vals])
            warnings.append(
                "no reporting bus assigned; columns C-H carried across from "
                "pf_external_grids (3P fault in amps, not kA)"
            )

        records.append(
            {
                "BSP": g["BSP_raw"],
                "Grid": g["Grid_raw"],
                "key": parse_grid(grid),
                "Buses": bus_text or (", ".join(m["candidates"][:12]) or None),
                "Source": (
                    f"{MAX_VALUE_TAB} / {MIN_VALUE_TAB}"
                    if buses
                    else "pf_external_grids (columns C-H carried across)"
                ),
                "d3P": (
                    None
                    if f3p is None or is_placeholder(f3p, rx) or not buses
                    else round(abs(float(f3p) / 1000.0 - float(vals_max[0])), 5)
                ),
                "Confidence": m["confidence"] or m["status"],
                "Note": "; ".join([n for n in ([m["note"]] + warnings) if n]),
                "assigned": tuple(buses),
                "candidates": tuple(m["candidates"]),
                "flag": m["status"] != "matched" or m["confidence"] != "High" or bool(warnings),
            }
        )

    flag_shared_buses(records)
    notes = [
        {k: r[k] for k in ("BSP", "Grid", "Buses", "Source", "d3P", "Confidence", "Note")}
        for r in records
        if r["flag"]
    ]

    write_workbook(template_path, out_path, results, notes)
    counts["rows"] = len(results)
    counts["grids"] = len(first)
    counts["flagged"] = len(notes)
    return counts


# --------------------------------------------------------------------------- #
# Output
# --------------------------------------------------------------------------- #

def write_workbook(template_path, out_path, results, notes) -> None:
    shutil.copy(template_path, out_path)
    wb = openpyxl.load_workbook(out_path)

    ws = wb["Grid Results"]
    body = Font(name="Arial", size=10)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None
    for r, values in enumerate(results, start=2):
        for c, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = body
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{len(results) + 1}"

    nt = wb["Mapping Notes"]
    for row in nt.iter_rows(min_row=1, max_row=nt.max_row, max_col=nt.max_column):
        for cell in row:
            cell.value = None
    headers = [
        "Bulk Supply Point",
        "Grid",
        "Mapped Bus(es) / candidates",
        "Source Tab(s)",
        "|dMax 3P fault| (kA)",
        "Confidence",
        "Note",
    ]
    head_font = Font(name="Arial", size=10, bold=True)
    for c, h in enumerate(headers, start=1):
        cell = nt.cell(row=1, column=c, value=h)
        cell.font = head_font
    for r, n in enumerate(notes, start=2):
        for c, key in enumerate(["BSP", "Grid", "Buses", "Source", "d3P", "Confidence", "Note"], start=1):
            cell = nt.cell(row=r, column=c, value=n[key])
            cell.font = body
            cell.alignment = Alignment(wrap_text=(key == "Note"), vertical="top")
    nt.column_dimensions["F"].width = 14
    nt.freeze_panes = "A2"
    nt.auto_filter.ref = f"A1:G{len(notes) + 1}"

    wb.save(out_path)


# --------------------------------------------------------------------------- #

def main() -> None:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--pf", required=True)
    p.add_argument("--report", required=True)
    p.add_argument("--template", required=True)
    p.add_argument("--out", required=True)
    a = p.parse_args()

    c = build(a.pf, a.report, a.template, a.out)
    print(
        f"{c['grids']} unique grids -> {c['rows']} rows\n"
        f"  matched   : {c['matched']}\n"
        f"  ambiguous : {c['ambiguous']}\n"
        f"  unmatched : {c['unmatched']}\n"
        f"  flagged in Mapping Notes: {c['flagged']}\n"
        f"written to {a.out}"
    )


if __name__ == "__main__":
    main()