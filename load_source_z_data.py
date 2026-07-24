"""
Module to build a grid data dictionary.

Reads published source-impedance studies from the ``Source Impedances``
folder in the repository root and returns a nested dictionary keyed by
external-grid name, for use by ``start.get_grid_data``.

Two sources are supported, selected by ``region``:

Regional Models (Ergon)
    ``2026 Fault Level Report (Ergon - Internal)_V1_1.xlsx``
    Keys come from column D ("Reported on Bus in PowerFactory") of the
    "Min Fault Level Report" tab. Maximum values come from the
    "Max-Max Fault Level Report" tab, minimum values from the
    "Min Fault Level Report" tab, and system-normal-minimum duplicates
    the minimum (the report publishes no separate system normal case).

SEQ (Energex)
    ``grid_results_all.xlsx``
    Keys come from column B ("Grid") of the "Grid Results" tab. Each
    grid has a "Max" row, a "Min" contingency row and a "Min" system
    normal row. Column F is published in amps and is converted to kA
    on read so that both regions return the same unit.

Returned structure (six elements per list: scenario name followed by
ikss, R/X, Z2/Z1, X0/X1, R0/X0)::

    {
        "ABPO_66kV_TEF T1": {
            "max":    ["System Normal", 2.03906, 0.40171, ...],
            "min":    ["System Normal", 1.54895, 0.53195, ...],
            "sn_min": ["System Normal", 1.54895, 0.53195, ...],
        },
        ...
    }

A missing workbook, sheet or key is never fatal: the function logs and
returns what it could build, leaving ``start.get_grid_data`` to fall
back to the values already stored on the PowerFactory model.
"""

import logging
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional, Tuple

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


# =============================================================================
# CONFIGURATION
# =============================================================================

# Folder holding the published source-impedance studies, relative to the
# repository root (the directory containing this module).
SOURCE_Z_DIR_NAME = "data"

# Preferred file names, with a glob fallback so that an annual re-issue
# (e.g. "2027 Fault Level Report ...") is still picked up. The resolved
# file is logged on every read so the provenance is in the run log.
REGIONAL_WORKBOOK = "2026 Fault Level Report (Ergon - Internal)_V1_1.xlsx"
REGIONAL_WORKBOOK_GLOB = "*Fault Level Report*.xlsx"
SEQ_WORKBOOK = "grid_results_all.xlsx"
SEQ_WORKBOOK_GLOB = "grid_results*.xlsx"

# Sheet names. Matching is whitespace/punctuation tolerant.
REGIONAL_MIN_SHEET = "Min Fault Level Report"
REGIONAL_MAX_SHEET = "Max-Max Fault Level Report"
SEQ_SHEET = "Grid Results"

# Regional report column layout (1-based).
_REG_KEY_COL = 4        # D - Reported on Bus in PowerFactory
_REG_FIRST_VAL_COL = 22  # V - PowerFactory short-circuit current
_REG_LAST_VAL_COL = 26   # Z - PowerFactory R0/X0 ratio

# SEQ workbook column layout (1-based).
_SEQ_KEY_COL = 2         # B - Grid, the outer dictionary key
_SEQ_BOUND_COL = 4       # D - Bound ("Max" / "Min")
_SEQ_FIRST_VAL_COL = 5   # E - Scenario
_SEQ_LAST_VAL_COL = 10   # J - R0/X1

SCENARIO_SYSTEM_NORMAL = "System Normal"
BOUND_MAX = "Max"
BOUND_MIN = "Min"

# ElmXnet holds ikss in kA. The Ergon report already publishes kA; the
# SEQ workbook publishes amps in column F, so it is converted on read
# and both regions return the same unit.
SEQ_IKSS_TO_KA = 0.001

# Both sources fall between 0.1 kA and 40 kA, so anything above this is
# not a kA figure and points at a change of units in the workbook.
_IKSS_KA_SANITY_LIMIT = 1000.0

# Parsed workbooks cached by (path, mtime, size) so an 80-project batch
# reads each file once rather than once per project.
_CACHE: Dict[Tuple[str, float, int], Dict[str, Dict[str, List[Any]]]] = {}


# =============================================================================
# PUBLIC API
# =============================================================================

def grid_data_import(
    region: str,
    source_dir: Optional[Path] = None,
) -> Dict[str, Dict[str, List[Any]]]:
    """
    Build the source-impedance dictionary for a region.

    Args:
        region: Network region string, as returned by
            ``pf_protection_helper.obtain_region``. 'Regional Models'
            selects the Ergon fault level report; anything else
            (i.e. 'SEQ') selects the Energex grid results workbook.
        source_dir: Optional override for the ``Source Impedances``
            folder. Defaults to the folder beside this module. Provided
            for testing and for pointing a run at an alternate study.

    Returns:
        Dict mapping external grid name to a dict with keys 'max',
        'min' and 'sn_min'. Each value is a six element list:
        ``[scenario, ikss, rntxn, z2tz1, x0tx1, r0tx0]``.
        Returns an empty dict if the workbook cannot be read.
    """
    directory = Path(source_dir) if source_dir else _default_source_dir()

    if region == 'Regional Models':
        workbook = _resolve_workbook(
            directory, REGIONAL_WORKBOOK, REGIONAL_WORKBOOK_GLOB
        )
        builder = _build_regional
    else:
        workbook = _resolve_workbook(
            directory, SEQ_WORKBOOK, SEQ_WORKBOOK_GLOB
        )
        builder = _build_seq

    if workbook is None:
        return {}

    cache_key = _cache_key(workbook)
    if cache_key in _CACHE:
        logger.info(
            "Source z-data for '%s' served from cache (%s entries): %s",
            region, len(_CACHE[cache_key]), workbook
        )
        return _CACHE[cache_key]

    logger.info("Reading source z-data for '%s' from %s", region, workbook)
    try:
        grid_data = builder(workbook)
    except Exception:
        logger.exception(
            "Failed to read source z-data workbook %s. "
            "Model default grid impedances will be kept.", workbook
        )
        return {}

    logger.info(
        "Loaded source z-data for %s external grids from %s",
        len(grid_data), workbook.name
    )
    _CACHE[cache_key] = grid_data
    return grid_data


def clear_cache() -> None:
    """Discard cached workbook contents (used by tests and reloads)."""
    _CACHE.clear()


# =============================================================================
# REGIONAL MODELS (ERGON)
# =============================================================================

def _build_regional(workbook: Path) -> Dict[str, Dict[str, List[Any]]]:
    """
    Build the grid data dictionary from the Ergon fault level report.

    Keys are taken from column D of the "Min Fault Level Report" tab in
    sheet order; the first occurrence of a duplicated key wins. 'max' is
    read from the matching column D row of the "Max-Max Fault Level
    Report" tab, 'min' from the row the key came from, and 'sn_min' is a
    copy of 'min'. A key with no "Max-Max" match is omitted entirely.

    Args:
        workbook: Path to the Ergon fault level report.

    Returns:
        The grid data dictionary.
    """
    grid_data: Dict[str, Dict[str, List[Any]]] = {}
    no_max: List[str] = []
    bad_values: List[str] = []

    wb = load_workbook(workbook, read_only=True, data_only=True)
    try:
        min_sheet = _resolve_sheet(wb, REGIONAL_MIN_SHEET, workbook)
        max_sheet = _resolve_sheet(wb, REGIONAL_MAX_SHEET, workbook)

        max_values = _first_match_by_key(max_sheet)

        for key, min_row in _iter_regional_rows(min_sheet):
            if key in grid_data:
                continue

            max_row = max_values.get(key)
            if max_row is None:
                no_max.append(key)
                continue

            if not (_all_numeric(min_row) and _all_numeric(max_row)):
                bad_values.append(key)
                continue

            minimum = [SCENARIO_SYSTEM_NORMAL] + list(min_row)
            grid_data[key] = {
                "max": [SCENARIO_SYSTEM_NORMAL] + list(max_row),
                "min": minimum,
                # 'sn_min' is a copy, not the same list object, so a
                # later mutation of one cannot silently alter the other.
                "sn_min": list(minimum),
            }
    finally:
        wb.close()

    _log_omissions(workbook, no_max, bad_values)
    _check_ikss_units(workbook, grid_data)
    return grid_data


def _iter_regional_rows(sheet) -> Iterator[Tuple[str, List[Any]]]:
    """
    Yield (key, values) for each data row of a fault level report tab.

    Args:
        sheet: An openpyxl worksheet for one of the report tabs.

    Yields:
        Tuples of the stripped column D key and the column V-Z values.
    """
    for row in sheet.iter_rows(
        min_row=2, max_col=_REG_LAST_VAL_COL, values_only=True
    ):
        key = _clean(_cell(row, _REG_KEY_COL))
        if not key:
            continue
        values = [
            _cell(row, col)
            for col in range(_REG_FIRST_VAL_COL, _REG_LAST_VAL_COL + 1)
        ]
        yield key, values


def _first_match_by_key(sheet) -> Dict[str, List[Any]]:
    """
    Index a fault level report tab by column D, keeping the first match.

    The "Max-Max Fault Level Report" tab repeats some bus names; the
    spec calls for the first matching row, so later rows are discarded.

    Args:
        sheet: An openpyxl worksheet for one of the report tabs.

    Returns:
        Dict mapping the column D key to its column V-Z values.
    """
    indexed: Dict[str, List[Any]] = {}
    for key, values in _iter_regional_rows(sheet):
        indexed.setdefault(key, values)
    return indexed


# =============================================================================
# SEQ (ENERGEX)
# =============================================================================

def _build_seq(workbook: Path) -> Dict[str, Dict[str, List[Any]]]:
    """
    Build the grid data dictionary from the Energex grid results.

    Grids are keyed by column B. Each contributes three rows: a "Max"
    row, a "Min" contingency row and a "Min" system normal row. Where a
    grid publishes no contingency minimum, 'min' falls back to the
    first "Min" row of any scenario. Fault currents are converted from
    amps to kA.

    Args:
        workbook: Path to grid_results_all.xlsx.

    Returns:
        The grid data dictionary.
    """
    grid_data: Dict[str, Dict[str, List[Any]]] = {}
    incomplete: List[str] = []
    bad_values: List[str] = []
    min_fallback: List[str] = []

    wb = load_workbook(workbook, read_only=True, data_only=True)
    try:
        sheet = _resolve_sheet(wb, SEQ_SHEET, workbook)
        by_grid = _group_seq_rows(sheet)

        for key, rows in by_grid.items():
            maximum = _select_seq_row(rows, BOUND_MAX)
            sn_minimum = _select_seq_row(
                rows, BOUND_MIN, scenario_is_system_normal=True
            )
            minimum = _select_seq_row(
                rows, BOUND_MIN, scenario_is_system_normal=False
            )

            if minimum is None:
                # 31 grids in the 2026 issue publish no contingency
                # row, only system normal minimums. The spec falls
                # back to any "Min" row in that case.
                minimum = _select_seq_row(rows, BOUND_MIN)
                if minimum is not None:
                    min_fallback.append(key)

            if maximum is None or minimum is None or sn_minimum is None:
                incomplete.append(key)
                continue

            if not all(
                _all_numeric(row[1:]) for row in (maximum, minimum, sn_minimum)
            ):
                bad_values.append(key)
                continue

            grid_data[key] = {
                "max": _to_ka(maximum),
                "min": _to_ka(minimum),
                "sn_min": _to_ka(sn_minimum),
            }
    finally:
        wb.close()

    if min_fallback:
        logger.info(
            "%s grids in %s publish no contingency minimum; the first "
            "'Min' row was used for 'min'. First few: %s",
            len(min_fallback), workbook.name, min_fallback[:5]
        )
    _log_omissions(workbook, incomplete, bad_values)
    _check_ikss_units(workbook, grid_data)
    return grid_data


def _group_seq_rows(sheet) -> Dict[str, List[List[Any]]]:
    """
    Group the "Grid Results" rows by their outer dictionary key.

    Args:
        sheet: The "Grid Results" worksheet.

    Returns:
        Dict mapping grid name to a list of its column E-J value rows,
        each prefixed by its stripped bound ("Max" or "Min").
    """
    by_grid: Dict[str, List[List[Any]]] = {}

    for row in sheet.iter_rows(
        min_row=2, max_col=_SEQ_LAST_VAL_COL, values_only=True
    ):
        key = _clean(_cell(row, _SEQ_KEY_COL))
        if not key:
            continue
        bound = _clean(_cell(row, _SEQ_BOUND_COL))
        values = [
            _cell(row, col)
            for col in range(_SEQ_FIRST_VAL_COL, _SEQ_LAST_VAL_COL + 1)
        ]
        # Scenario is a label, not a measurement; strip it here so
        # comparisons and the returned list agree.
        values[0] = _clean(values[0])
        by_grid.setdefault(key, []).append([bound] + values)

    return by_grid


def _select_seq_row(
    rows: List[List[Any]],
    bound: str,
    scenario_is_system_normal: Optional[bool] = None,
) -> Optional[List[Any]]:
    """
    Return the first row matching a bound and optional scenario test.

    Args:
        rows: Candidate rows for one grid, each ``[bound, scenario, ...]``.
        bound: "Max" or "Min".
        scenario_is_system_normal: None to ignore the scenario, True to
            require "System Normal", False to require anything else.

    Returns:
        The six element ``[scenario, ikss, ...]`` list, or None.
    """
    for row in rows:
        if row[0] != bound:
            continue
        if scenario_is_system_normal is not None:
            is_sn = row[1] == SCENARIO_SYSTEM_NORMAL
            if is_sn != scenario_is_system_normal:
                continue
        return list(row[1:])
    return None


def _to_ka(row: List[Any]) -> List[Any]:
    """
    Convert the fault current element of a SEQ row from amps to kA.

    Only element 1 is a current; the remaining elements are ratios and
    are left alone.

    Args:
        row: A six element ``[scenario, ikss, ...]`` list.

    Returns:
        A new list with element 1 in kA.
    """
    converted = list(row)
    converted[1] = converted[1] * SEQ_IKSS_TO_KA
    return converted


# =============================================================================
# FILE AND SHEET RESOLUTION
# =============================================================================

def _default_source_dir() -> Path:
    """Return the ``Source Impedances`` folder beside this module."""
    return Path(__file__).resolve().parent / SOURCE_Z_DIR_NAME


def _resolve_workbook(
    directory: Path,
    preferred_name: str,
    pattern: str,
) -> Optional[Path]:
    """
    Locate a source-impedance workbook.

    Tries the exact expected name first, then the glob pattern so an
    annual re-issue under a new name is still found. Temporary Excel
    lock files (``~$``) are ignored.

    Args:
        directory: The ``Source Impedances`` folder.
        preferred_name: Exact file name from the specification.
        pattern: Glob fallback pattern.

    Returns:
        Path to the workbook, or None if nothing usable was found.
    """
    if not directory.is_dir():
        logger.error(
            "Source impedance folder not found: %s. Model default grid "
            "impedances will be kept for every grid.", directory
        )
        return None

    exact = directory / preferred_name
    if exact.is_file():
        return exact

    candidates = sorted(
        path for path in directory.glob(pattern)
        if path.is_file() and not path.name.startswith("~$")
    )
    if not candidates:
        logger.error(
            "No source impedance workbook matching '%s' or '%s' in %s. "
            "Model default grid impedances will be kept for every grid.",
            preferred_name, pattern, directory
        )
        return None

    chosen = candidates[-1]
    logger.warning(
        "Expected source impedance workbook '%s' not found in %s; "
        "using '%s' instead.", preferred_name, directory, chosen.name
    )
    return chosen


def _resolve_sheet(workbook, sheet_name: str, path: Path):
    """
    Return a worksheet, tolerating punctuation and spacing differences.

    Args:
        workbook: An open openpyxl workbook.
        sheet_name: The expected sheet name.
        path: Workbook path, for the error message.

    Returns:
        The matching worksheet.

    Raises:
        KeyError: If no sheet matches.
    """
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    target = _normalise(sheet_name)
    for name in workbook.sheetnames:
        if _normalise(name) == target:
            logger.warning(
                "Sheet '%s' not found in %s; matched '%s' instead.",
                sheet_name, path.name, name
            )
            return workbook[name]

    raise KeyError(
        f"Sheet '{sheet_name}' not found in {path.name}. "
        f"Available sheets: {workbook.sheetnames}"
    )


def _cache_key(path: Path) -> Tuple[str, float, int]:
    """Return a cache key that changes when the workbook changes."""
    stat = path.stat()
    return (str(path), stat.st_mtime, stat.st_size)


# =============================================================================
# HELPERS
# =============================================================================

def _cell(row: Tuple[Any, ...], column: int) -> Any:
    """
    Return a 1-based column from a ``values_only`` row tuple.

    Args:
        row: Row tuple from ``iter_rows(values_only=True)``.
        column: 1-based column number.

    Returns:
        The cell value, or None if the row is short.
    """
    index = column - 1
    return row[index] if index < len(row) else None


def _clean(value: Any) -> str:
    """Return a stripped string for a cell value, '' for blanks."""
    return "" if value is None else str(value).strip()


def _normalise(name: str) -> str:
    """Lower-case a name and drop everything but alphanumerics."""
    return "".join(char for char in name.lower() if char.isalnum())


def _all_numeric(values) -> bool:
    """Return True if every value is a real number (bools excluded)."""
    return all(
        isinstance(value, (int, float)) and not isinstance(value, bool)
        for value in values
    )


def _log_omissions(
    workbook: Path,
    missing: List[str],
    bad_values: List[str],
) -> None:
    """
    Log grids left out of the dictionary, with a sample of the names.

    Args:
        workbook: Workbook the data came from.
        missing: Keys with no complete set of rows.
        bad_values: Keys whose values were blank or non-numeric.
    """
    if missing:
        logger.warning(
            "%s grids in %s had no complete max/min row set and were "
            "omitted; model defaults will be kept for these. "
            "First few: %s",
            len(missing), workbook.name, missing[:5]
        )
    if bad_values:
        logger.warning(
            "%s grids in %s had blank or non-numeric impedance values "
            "and were omitted; model defaults will be kept for these. "
            "First few: %s",
            len(bad_values), workbook.name, bad_values[:5]
        )


def _check_ikss_units(
    workbook: Path,
    grid_data: Dict[str, Dict[str, List[Any]]],
) -> None:
    """
    Warn when fault currents look wrong for the ElmXnet ikss unit (kA).

    ``start.get_grid_data`` writes element 1 of each list straight to
    ``ikss``/``ikssmin``, which PowerFactory holds in kA. Values in the
    thousands are amps and would overstate the source by 1000x; values
    at or below zero cannot be applied to an in-service grid.

    Args:
        workbook: Workbook the data came from.
        grid_data: The dictionary just built.
    """
    amps_like: List[str] = []
    non_positive: List[str] = []

    for key, scenarios in grid_data.items():
        currents = [scenarios[name][1] for name in ("max", "min", "sn_min")]
        if any(value > _IKSS_KA_SANITY_LIMIT for value in currents):
            amps_like.append(key)
        if any(value <= 0 for value in currents):
            non_positive.append(key)

    if amps_like:
        logger.error(
            "%s of %s grids in %s have fault currents above %s kA. "
            "ikss is set in kA, so the workbook units have probably "
            "changed and every value would be 1000x too high. Check "
            "the source before trusting this run. First few: %s",
            len(amps_like), len(grid_data), workbook.name,
            _IKSS_KA_SANITY_LIMIT, amps_like[:5]
        )
    if non_positive:
        logger.warning(
            "%s grids in %s have a fault current of zero or less; these "
            "grids cannot supply fault current in the study. First "
            "few: %s",
            len(non_positive), workbook.name, non_positive[:5]
        )